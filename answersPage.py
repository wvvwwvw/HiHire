import json
import smtplib
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (QFrame, QVBoxLayout, QLabel, QPushButton, QWidget,
                               QHBoxLayout, QScrollArea, QMessageBox, QFileDialog)
from PySide6.QtCore import Qt, Signal, QDate, QSize
from connect_to_database import Database
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


class AnswersPage:
    def __init__(self, ui, main_window):
        self.ui = ui
        self.main_window = main_window
        self.db = Database()
        self.current_interview_id = None
        self.current_user_role = None
        self.current_user_id = None

        self.connect_buttons()
        self.connect_filters()

    def connect_buttons(self):
        """Подключает обработчики кнопок"""
        self.ui.exitFromCandidateBtn.clicked.connect(self.return_to_reports_page)
        self.ui.exportBtn.clicked.connect(self.export_interview_to_excel)
        self.ui.importBtn.clicked.connect(self.import_interview_from_excel)

    def connect_filters(self):
        """Подключает обработчики фильтров"""
        self.ui.statusesComboBox.currentIndexChanged.connect(self.filter_candidates)
        self.ui.positionComboBox.currentIndexChanged.connect(self.filter_candidates)
        self.ui.candidateLineEdit.textChanged.connect(self.filter_candidates)
        self.ui.scoreSortComboBox.currentIndexChanged.connect(self.filter_candidates)
        self.ui.startDateEdit.dateChanged.connect(self.filter_candidates)
        self.ui.endDateEdit.dateChanged.connect(self.filter_candidates)

    def init_reports_page(self, role_id, user_id):
        """Инициализирует страницу отчетов"""
        self.current_user_role = role_id
        self.current_user_id = user_id

        self.load_statuses()
        self.load_positions()
        self.setup_date_filters()
        self.setup_score_sorting()
        self.load_candidates()

    def load_statuses(self):
        """Загружает статусы кандидатов"""
        statuses = self.db.execute_query("SELECT id, name FROM Statuses WHERE type = 'cand'")
        self.ui.statusesComboBox.clear()
        self.ui.statusesComboBox.addItem("Все статусы", None)
        for status in statuses:
            self.ui.statusesComboBox.addItem(status['name'], status['id'])

    def load_positions(self):
        """Загружает список должностей"""
        positions = self.db.execute_query("SELECT id, name FROM Positions")
        self.ui.positionComboBox.clear()
        self.ui.positionComboBox.addItem("Все должности", None)
        for position in positions:
            self.ui.positionComboBox.addItem(position['name'], position['id'])

    def setup_date_filters(self):
        """Настраивает фильтры по дате"""
        today = QDate.currentDate()
        self.ui.endDateEdit.setDate(today)

    def setup_score_sorting(self):
        """Настраивает сортировку по баллам"""
        self.ui.scoreSortComboBox.clear()
        self.ui.scoreSortComboBox.addItem("Без сортировки", None)
        self.ui.scoreSortComboBox.addItem("По возрастанию баллов", "asc")
        self.ui.scoreSortComboBox.addItem("По убыванию баллов", "desc")

    def load_candidates(self, filters=None):
        """Загружает кандидатов с учетом фильтров"""
        self.clear_layout(self.ui.candidatesVerticalLayout)
        self.ui.candidatesVerticalLayout.setAlignment(Qt.AlignTop)

        query, params = self.build_candidates_query(filters)
        interviews = self.db.execute_query(query, params or None)

        if not interviews:
            self.show_no_results_message()
            return

        for interview in interviews:
            self.add_candidate_card(interview)

    def build_candidates_query(self, filters):
        """Строит SQL запрос для получения кандидатов"""
        query = """
        SELECT i.id, i.questionnaire_id, i.candidate_id, i.status_id, i.created_at,
               q.title as questionnaire_title,
               c.first_name, c.last_name, c.patronimic, c.resume_link, c.email,
               s.name as status_name,
               p.name as position_name,
               SUM(a.value_numeric) as total_score
        FROM Interviews i
        JOIN Questionnaires q ON i.questionnaire_id = q.id
        JOIN Candidates c ON i.candidate_id = c.id
        JOIN Statuses s ON i.status_id = s.id
        JOIN Positions p ON q.position_id = p.id
        LEFT JOIN Answers a ON i.id = a.interview_id
        WHERE 1=1
        """
        params = []

        if self.current_user_role != 1:
            query += " AND i.interviewer_id = %s"
            params.append(self.current_user_id)

        if filters:
            query, params = self.apply_filters(query, params, filters)

        query += " GROUP BY i.id, i.questionnaire_id, i.candidate_id, i.status_id, i.created_at, "
        query += "q.title, c.first_name, c.last_name, c.patronimic, c.resume_link, s.name, p.name"

        if filters and filters.get('score_sort'):
            query += " ORDER BY total_score " + filters['score_sort']
        else:
            query += " ORDER BY i.id DESC"

        return query, params

    def apply_filters(self, query, params, filters):
        """Добавляет условия фильтрации в SQL запрос"""
        if filters.get('status_id'):
            query += " AND i.status_id = %s"
            params.append(filters['status_id'])

        if filters.get('position_id'):
            query += " AND q.position_id = %s"
            params.append(filters['position_id'])

        if filters.get('candidate_name'):
            query += " AND (c.first_name LIKE %s OR c.last_name LIKE %s OR c.patronimic LIKE %s)"
            params.extend([f'%{filters["candidate_name"]}%'] * 3)

        if filters.get('start_date'):
            query += " AND DATE(i.created_at) >= %s"
            params.append(filters['start_date'])

        if filters.get('end_date'):
            query += " AND DATE(i.created_at) <= %s"
            params.append(filters['end_date'])

        return query, params

    def show_no_results_message(self):
        """Показывает сообщение об отсутствии результатов"""
        no_results_label = QLabel("Нет интервью, соответствующих фильтрам")
        no_results_label.setStyleSheet(f"font-size: 12pt; color: {self.main_window.theme.COLOR_TEXT_2};")
        self.ui.candidatesVerticalLayout.addWidget(no_results_label)

    def add_candidate_card(self, interview):
        """Добавляет карточку кандидата"""
        frame = QFrame()
        frame.setMaximumHeight(130)
        frame.setFrameShape(QFrame.StyledPanel)
        frame.setFrameShadow(QFrame.Raised)
        frame.setStyleSheet(self.get_card_style())

        layout = QHBoxLayout(frame)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        self.add_left_card_content(layout, interview)
        self.add_center_card_content(layout, interview)

        if self.current_user_role == 1:
            self.add_right_card_content(layout, interview)

        frame.mousePressEvent = lambda _, i=interview: self.open_candidate_page(i)
        self.ui.candidatesVerticalLayout.addWidget(frame)

    def get_card_style(self):
        """Возвращает стиль для карточки кандидата"""
        return f"""
            QFrame {{
                background-color: {self.main_window.theme.COLOR_BACKGROUND_1};
                border-radius: 10px;
                margin: 2px;
            }}
            QFrame:hover {{
                background-color: {self.main_window.theme.COLOR_BACKGROUND_3};
                border: {self.main_window.theme.COLOR_ACCENT_1};
            }}
        """

    def add_left_card_content(self, layout, interview):
        """Добавляет левую часть карточки (информация о кандидате)"""
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(5)

        questionnaire_label = QLabel(interview['questionnaire_title'])
        questionnaire_label.setStyleSheet(self.get_label_style(12, True))
        questionnaire_label.setFixedHeight(35)
        left_layout.addWidget(questionnaire_label)

        fio = f"{interview['last_name']} {interview['first_name']} {interview['patronimic'] or ''}"
        fio_label = QLabel(fio.strip())
        fio_label.setStyleSheet(self.get_label_style(11))
        fio_label.setFixedHeight(30)
        left_layout.addWidget(fio_label)

        if interview['resume_link']:
            resume_label = QLabel(f"Резюме: {interview['resume_link']}")
            resume_label.setStyleSheet("font-size: 10pt; color: #1E90FF; background-color: transparent;")
            resume_label.setOpenExternalLinks(True)
            resume_label.setFixedHeight(25)
            left_layout.addWidget(resume_label)

        layout.addWidget(left_widget, 60)

    def add_center_card_content(self, layout, interview):
        """Добавляет центральную часть карточки (статус и баллы)"""
        center_widget = QWidget()
        center_layout = QVBoxLayout(center_widget)
        center_layout.setAlignment(Qt.AlignRight | Qt.AlignTop)
        center_layout.setContentsMargins(0, 0, 0, 0)
        center_layout.setSpacing(5)

        status_label = QLabel(interview['status_name'])
        status_label.setStyleSheet(self.get_status_label_style())
        status_label.setFixedHeight(25)
        status_label.setAlignment(Qt.AlignCenter)
        center_layout.addWidget(status_label)

        max_score = self.db.get_max_score_for_questionnaire(interview['questionnaire_id'])
        score_label = QLabel(f"Баллы: {interview['total_score'] or 0} из {max_score}")
        score_label.setStyleSheet(self.get_label_style(10))
        score_label.setFixedHeight(25)
        score_label.setAlignment(Qt.AlignRight)
        center_layout.addWidget(score_label)

        date_label = QLabel(interview['created_at'].strftime('%d.%m.%Y'))
        date_label.setStyleSheet(self.get_label_style(10))
        date_label.setFixedHeight(25)
        date_label.setAlignment(Qt.AlignRight)
        center_layout.addWidget(date_label)

        layout.addWidget(center_widget, 30)

    def add_right_card_content(self, layout, interview):
        """Добавляет правую часть карточки (кнопка удаления)"""
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setAlignment(Qt.AlignRight | Qt.AlignTop)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(5)

        delete_btn = QPushButton("-")
        delete_btn.setFixedSize(100, 20)
        delete_btn.setStyleSheet(self.get_button_style())
        delete_btn.clicked.connect(lambda: self.delete_interview(interview))
        right_layout.addWidget(delete_btn)

        layout.addWidget(right_widget, 10)

    def get_label_style(self, size, bold=False):
        """Возвращает стиль для текстовых меток"""
        weight = "font-weight: bold;" if bold else ""
        return f"""
            {weight} font-size: {size}pt; 
            background-color: transparent; 
            color: {self.main_window.theme.COLOR_TEXT_1};
        """

    def get_status_label_style(self):
        """Возвращает стиль для метки статуса"""
        return f"""
            font-size: 11pt; 
            font-weight: bold; 
            color: {self.main_window.theme.COLOR_TEXT_1};
            background-color: {self.main_window.theme.COLOR_ACCENT_1};
            border-radius: 10px;
            padding: 2px 5px;
        """

    def get_button_style(self):
        """Возвращает стиль для кнопок"""
        return f"""
            QPushButton {{
                background-color: {self.main_window.theme.COLOR_ACCENT_1};
                color: {self.main_window.theme.COLOR_TEXT_1};
                border-radius: 10px;
                font-size: 10pt;
            }}
            QPushButton:hover {{
                background-color: {self.main_window.theme.COLOR_ACCENT_2};
            }}
        """

    def delete_interview(self, interview):
        """Удаляет интервью, связанного кандидата и все ответы с подтверждением"""
        confirm_dialog = QMessageBox(self.main_window)
        confirm_dialog.setWindowTitle("Подтверждение удаления")
        confirm_dialog.setText(
            f"Вы уверены, что хотите удалить кандидата {interview['last_name']} {interview['first_name']} "
            "и все связанные данные?"
        )

        confirm_dialog.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        confirm_dialog.setButtonText(QMessageBox.Yes, "Да")
        confirm_dialog.setButtonText(QMessageBox.No, "Нет")

        confirm_dialog.setDefaultButton(QMessageBox.No)
        confirm_dialog.setIcon(QMessageBox.Warning)

        confirm_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {self.main_window.theme.COLOR_BACKGROUND_1};
                color: {self.main_window.theme.COLOR_TEXT_1};
            }}
            QMessageBox QLabel {{
                color: {self.main_window.theme.COLOR_TEXT_1};
            }}
        """)

        if confirm_dialog.exec() == QMessageBox.Yes:
            try:
                self.db.execute_query(
                    "DELETE FROM Answers WHERE interview_id = %s",
                    (interview['id'],))

                self.db.execute_query(
                    "DELETE FROM Interviews WHERE id = %s",
                    (interview['id'],))

                self.db.execute_query(
                    "DELETE FROM Emails WHERE candidate_id = %s",
                    (interview['candidate_id'],))

                self.db.execute_query(
                    "DELETE FROM Candidates WHERE id = %s",
                    (interview['candidate_id'],))

                self.load_candidates()

                QMessageBox.information(
                    self.main_window,
                    "Удаление завершено",
                    "Кандидат и все связанные данные успешно удалены.",
                    QMessageBox.Ok
                )

            except Exception as e:
                print(f"Ошибка при удалении интервью и связанных данных: {e}")
                # Показываем сообщение об ошибке
                QMessageBox.critical(
                    self.main_window,
                    "Ошибка удаления",
                    f"Произошла ошибка при удалении: {str(e)}",
                    QMessageBox.Ok
                )

    def open_candidate_page(self, interview):
        """Открывает страницу кандидата с подробной информацией"""
        self.current_interview_id = interview['id']

        fio = f"{interview['last_name']} {interview['first_name']} {interview['patronimic'] or ''}"
        self.ui.candidateFIOLabel.setText(fio.strip())

        self.ui.positionName.setText(interview['position_name'])

        current_status = interview['status_name']
        self.ui.statusLabel.setText(current_status)
        self.ui.statusLabel.setStyleSheet(f"""
            font-size: 11pt; 
            font-weight: bold; 
            color: {self.main_window.theme.COLOR_TEXT_1};
            background-color: {self.main_window.theme.COLOR_ACCENT_1};
            border-radius: 10px;
            padding: 2px 5px;
        """)

        max_score = self.db.get_max_score_for_questionnaire(interview['questionnaire_id'])
        total_score = interview['total_score'] or 0
        self.ui.scoreLabel.setText(f"{total_score} из {max_score}")

        self.ui.questionnaryName.setText(interview['questionnaire_title'])

        if interview['resume_link']:
            self.ui.linkLabel.setText(f'<a href="{interview["resume_link"]}">{interview["resume_link"]}</a>')
            self.ui.linkLabel.setOpenExternalLinks(True)
        else:
            self.ui.linkLabel.setText("Не указано")

        self.ui.emailLabel.setText(interview['email'])

        self.load_answers(interview['id'], interview['questionnaire_id'])

        category_scores = self.get_category_scores(interview['id'])

        category_text = "\n".join([f"{category}: {score} баллов" for category, score in category_scores.items()])

        percentage = (total_score / max_score) * 100 if max_score else 0
        if percentage < 50:
            recommendation = "Низкий результат. Кандидат не соответствует требованиям."
        elif percentage < 80:
            recommendation = "Средний результат. Кандидат соответствует минимальным требованиям."
        else:
            recommendation = "Высокий результат. Кандидат полностью соответствует требованиям."

        full_recommendation = f"Рекомендация: {recommendation}\n\nБаллы по категориям:\n{category_text}"
        self.ui.recommendationText.setText(full_recommendation)

        self.setup_buttons_based_on_status(current_status)

        self.connect_buttons_handlers(interview)

        self.ui.mainPages.setCurrentWidget(self.ui.candidatePage)

    def get_category_scores(self, interview_id):
        """Возвращает баллы кандидата по категориям вопросов"""
        query = """
        SELECT 
            c.name as category_name,
            SUM(a.value_numeric) as category_score
        FROM Answers a
        JOIN Questions q ON a.question_id = q.id
        JOIN Categories c ON q.category_id = c.id
        WHERE a.interview_id = %s
        GROUP BY c.name
        """
        results = self.db.execute_query(query, (interview_id,))

        category_scores = {
            "Технические знания": 0,
            "Коммуникация": 0,
            "Инструменты": 0,
            "Практические навыки": 0
        }

        for result in results:
            category_name = result['category_name']
            if category_name in category_scores:
                category_scores[category_name] = result['category_score'] or 0

        return category_scores

    def setup_buttons_based_on_status(self, current_status):
        """Настраивает видимость кнопок в зависимости от статуса кандидата"""
        self.ui.acceptBtn.hide()
        self.ui.rejectBtn.hide()
        self.ui.disagreeBtn.hide()
        self.ui.agreeBtn.hide()

        if current_status == "Новый":
            self.ui.acceptBtn.show()
            self.ui.rejectBtn.show()
        elif current_status == "Ожидание ответа":
            self.ui.disagreeBtn.show()
            self.ui.agreeBtn.show()

    def connect_buttons_handlers(self, interview):
        """Подключает обработчики для кнопок действий с кандидатом"""
        try:
            self.ui.rejectBtn.clicked.disconnect()
            self.ui.acceptBtn.clicked.disconnect()
            self.ui.disagreeBtn.clicked.disconnect()
            self.ui.agreeBtn.clicked.disconnect()
        except:
            pass

        user_email = self.get_user_email(self.current_user_id)

        self.ui.rejectBtn.clicked.connect(
            lambda: self.send_status_email(interview, user_email, status_id=3, is_rejection=True)
        )
        self.ui.acceptBtn.clicked.connect(
            lambda: self.send_status_email(interview, user_email, status_id=7)  # Ожидание ответа
        )
        self.ui.disagreeBtn.clicked.connect(
            lambda: self.update_candidate_status(interview['candidate_id'], 3)  # Отказ
        )
        self.ui.agreeBtn.clicked.connect(lambda: self.hire_candidate(interview))

    def send_status_email(self, interview, sender_email, status_id, is_rejection=False):
        """Отправляет письмо кандидату в зависимости от статуса"""
        try:
            candidate = self.db.execute_query(
                "SELECT email FROM Candidates WHERE id = %s",
                (interview['candidate_id'],),
                fetch_all=False
            )

            if not candidate or not candidate['email']:
                QMessageBox.warning(self.main_window, "Ошибка", "У кандидата не указан email")
                return

            user_data = self.db.execute_query(
                "SELECT work_email_password FROM Users WHERE id = %s",
                (self.current_user_id,),
                fetch_all=False
            )

            if not user_data or not user_data['work_email_password']:
                QMessageBox.warning(self.main_window, "Ошибка", "Не удалось получить пароль для почты")
                return

            if is_rejection:
                subject = "Результаты собеседования"
                message = f"""
                Уважаемый кандидат,\n\n
                К сожалению, мы не можем предложить вам позицию {interview['position_name']}.\n\n
                Спасибо за ваше время и интерес к нашей компании.\n\n
                С уважением,\n
                Команда HR,\n
                Большая Тройка
                """
                email_status_id = 9
            else:
                subject = "Приглашение на работу"
                message = f"""
                Уважаемый кандидат,\n\n
                Мы рады пригласить Вас в свою команду. Просим дать обратную связь в течении 7 дней.\n\n
                Должность: {interview['position_name']}\n
                С уважением,\n
                Команда HR, \n
                Большая Тройка
                """
                email_status_id = 8

            if self.send_email(
                    sender_email,
                    user_data['work_email_password'],
                    candidate['email'],
                    subject,
                    message
            ):
                self.db.execute_query(
                    "INSERT INTO Emails (sender_id, candidate_id, subject, message, status_id) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (self.current_user_id, interview['candidate_id'], subject, message, email_status_id))

                self.update_candidate_status(interview['candidate_id'], status_id)

                QMessageBox.information(
                    self.main_window,
                    "Успех",
                    "Письмо успешно отправлено" if is_rejection else "Письмо-приглашение успешно отправлено"
                )
            else:
                QMessageBox.warning(self.main_window, "Ошибка", "Не удалось отправить письмо")

        except Exception as e:
            QMessageBox.critical(self.main_window, "Ошибка", f"Произошла ошибка: {str(e)}")

    def get_user_email(self, user_id):
        """Получает email пользователя из связанной таблицы Employees"""
        try:
            query = """
            SELECT e.email 
            FROM Users u
            JOIN Employees e ON u.employee_id = e.id
            WHERE u.id = %s
            """
            result = self.db.execute_query(query, (user_id,), fetch_all=False)
            return result['email'] if result else None
        except Exception as e:
            print(f"Ошибка при получении email пользователя: {e}")
            return None

    def update_candidate_status(self, candidate_id, new_status_id):
        """Обновляет статус кандидата"""
        try:
            self.db.execute_query(
                "UPDATE Candidates SET status_id = %s WHERE id = %s",
                (new_status_id, candidate_id))

            self.db.execute_query("UPDATE Interviews SET status_id = %s WHERE candidate_id = %s",
                                  (new_status_id, candidate_id))

            self.return_to_reports_page()
            self.load_candidates()
        except Exception as e:
            print(f"Ошибка при обновлении статуса кандидата: {e}")

    def send_email(self, sender_email, sender_password, recipient_email, subject, message):
        """Отправляет email через SMTP"""
        try:
            if 'yandex' in sender_email.lower() or 'ya' in sender_email.lower():
                smtp_server = 'smtp.yandex.ru'
                smtp_port = 465
            elif 'mail' in sender_email.lower():
                smtp_server = 'smtp.mail.ru'
                smtp_port = 465
            elif 'gmail' in sender_email.lower():
                smtp_server = 'smtp.gmail.com'
                smtp_port = 587
            else:
                smtp_server = 'smtp.yandex.ru'
                smtp_port = 465

            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = recipient_email
            msg['Subject'] = subject

            msg.attach(MIMEText(message, 'plain'))

            with smtplib.SMTP_SSL(smtp_server, smtp_port) if smtp_port == 465 else smtplib.SMTP(smtp_server,
                                                                                                smtp_port) as server:
                if smtp_port == 587:
                    server.starttls()

                server.login(sender_email, sender_password)
                server.send_message(msg)

            return True
        except Exception as e:
            print(f"Ошибка при отправке письма: {e}")
            return False

    def hire_candidate(self, interview):
        """Создает сотрудника на основе кандидата и обновляет статус"""
        try:
            questionnaire = self.db.execute_query(
                "SELECT position_id FROM Questionnaires WHERE id = %s",
                (interview['questionnaire_id'],), fetch_all=False)

            if not questionnaire:
                print("Ошибка: не найдена анкета для интервью")
                return

            position_id = questionnaire['position_id']

            answers = self.db.execute_query(
                "SELECT a.question_id, a.value_numeric, a.selected_options, q.category_id, q.answer_type_id "
                "FROM Answers a "
                "JOIN Questions q ON a.question_id = q.id "
                "WHERE a.interview_id = %s",
                (interview['id'],))

            skills_by_category = {
                1: [],  # Технические знания
                2: [],  # Коммуникация
                3: [],  # Инструменты
                4: []  # Практические навыки
            }

            for answer in answers:
                category_id = answer['category_id']
                answer_type_id = answer['answer_type_id']

                if answer_type_id == 5:
                    continue

                if answer_type_id in (1, 2):
                    selected_options = answer.get('selected_options', '[]')
                    if selected_options and isinstance(selected_options, str):
                        try:
                            selected_ids = eval(selected_options)
                            if isinstance(selected_ids, list):
                                options = self.db.execute_query(
                                    "SELECT text FROM QuestionOptions WHERE id IN %s",
                                    (selected_ids,))
                                for opt in options:
                                    skills_by_category[category_id].append(opt['text'])
                        except:
                            pass
                elif answer_type_id in (3, 4, 6):
                    if answer['value_numeric'] is not None:
                        skills_by_category[category_id].append(str(answer['value_numeric']))
                    elif answer.get('selected_options'):
                        skills_by_category[category_id].append(answer['selected_options'])

            skills = {
                "Технические знания": ", ".join(skills_by_category[1]),
                "Коммуникация": ", ".join(skills_by_category[2]),
                "Инструменты": ", ".join(skills_by_category[3]),
                "Практические навыки": ", ".join(skills_by_category[4])
            }

            candidate = self.db.execute_query(
                "SELECT first_name, last_name, patronimic, email, phone FROM Candidates WHERE id = %s",
                (interview['candidate_id'],), fetch_all=False)

            if candidate:
                self.db.execute_query(
                    "INSERT INTO Employees (first_name, last_name, patronimic, email, phone, "
                    "position_id, hire_date, status_id, skills, organization_id) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (
                        candidate['first_name'],
                        candidate['last_name'],
                        candidate['patronimic'],
                        candidate['email'],
                        candidate['phone'],
                        position_id,
                        QDate.currentDate().toString("yyyy-MM-dd"),
                        10,
                        str(skills),
                        1
                    ))

                self.update_candidate_status(interview['candidate_id'], 4)
        except Exception as e:
            print(f"Ошибка при найме кандидата: {e}")

    def load_answers(self, interview_id, questionnaire_id):
        """Загружает ответы кандидата"""
        self.clear_layout(self.ui.answersWidget.layout())

        query = """
        SELECT q.id, q.text, q.weight, q.answer_type_id, q.category_id
        FROM Questions q
        JOIN QuestionnaireQuestions qq ON q.id = qq.question_id
        WHERE qq.questionnaire_id = %s
        """
        questions = self.db.execute_query(query, (questionnaire_id,))

        answers = self.db.execute_query(
            "SELECT question_id, value_numeric, selected_options, text_answer FROM Answers WHERE interview_id = %s",
            (interview_id,)
        )

        answers_dict = {a['question_id']: a for a in answers}

        for question in questions:
            self.add_answer_to_layout(question, answers_dict.get(question['id']))

    def add_answer_to_layout(self, question, answer):
        """Добавляет вопрос и ответ в layout"""
        answer_widget = QWidget()
        layout = QVBoxLayout(answer_widget)
        layout.setSpacing(5)
        layout.setContentsMargins(0, 0, 0, 0)

        question_label = QLabel(question['text'])
        question_label.setWordWrap(True)
        question_label.setStyleSheet("font-weight: bold; font-size: 12pt;")
        layout.addWidget(question_label)

        if answer:
            if question['answer_type_id'] == 1:
                selected_ids = answer.get('selected_options', '[]')
                if selected_ids and isinstance(selected_ids, str):
                    try:
                        selected_ids = eval(selected_ids)
                    except:
                        selected_ids = []

                options = self.db.get_options_for_question(question['id'])
                selected_options = [opt['text'] for opt in options if opt['id'] in selected_ids]
                answer_text = ", ".join(selected_options) if selected_options else "Нет ответа"

            elif question['answer_type_id'] == 2:
                selected_id = answer.get('selected_options', '[]')
                if selected_id and isinstance(selected_id, str):
                    try:
                        selected_id = eval(selected_id)
                        selected_id = selected_id[0] if selected_id else None
                    except:
                        selected_id = None

                option = self.db.execute_query(
                    "SELECT text FROM QuestionOptions WHERE id = %s",
                    (selected_id,),
                    fetch_all=False
                ) if selected_id else None

                answer_text = option['text'] if option else "Нет ответа"

            elif question['answer_type_id'] in (3, 4):
                answer_text = str(answer.get('value_numeric', 'Нет ответа'))

            elif question['answer_type_id'] == 5:
                answer_value = answer.get('selected_options', '[]')
                if answer_value and isinstance(answer_value, str):
                    try:
                        answer_value = eval(answer_value)
                        answer_value = answer_value[0] if answer_value else None
                    except:
                        answer_value = None

                answer_text = answer_value if answer_value else "Нет ответа"

            elif question['answer_type_id'] == 6:
                answer_text = answer.get('text_answer', 'Нет ответа')

            answer_label = QLabel(f"Ответ: {answer_text}")
            answer_label.setWordWrap(True)
            answer_label.setStyleSheet("font-size: 12pt; margin-left: 10px;")
            layout.addWidget(answer_label)

        self.ui.answersWidget.layout().addWidget(answer_widget)

    def filter_candidates(self):
        """Применяет фильтры к списку кандидатов"""
        filters = {}

        status_data = self.ui.statusesComboBox.currentData()
        if status_data is not None:
            filters['status_id'] = status_data

        position_data = self.ui.positionComboBox.currentData()
        if position_data is not None:
            filters['position_id'] = position_data

        candidate_name = self.ui.candidateLineEdit.text().strip()
        if candidate_name:
            filters['candidate_name'] = candidate_name

        score_sort = self.ui.scoreSortComboBox.currentData()
        if score_sort:
            filters['score_sort'] = score_sort

        start_date = self.ui.startDateEdit.date()
        if not start_date.isNull():
            filters['start_date'] = start_date.toString("yyyy-MM-dd")

        end_date = self.ui.endDateEdit.date()
        if not end_date.isNull():
            filters['end_date'] = end_date.toString("yyyy-MM-dd")

        self.load_candidates(filters)

    def return_to_reports_page(self):
        """Возвращает на страницу отчетов"""
        self.ui.mainPages.setCurrentWidget(self.ui.reportsPage)

    def clear_layout(self, layout):
        """Очищает layout от всех виджетов"""
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())

    def export_interview_to_excel(self):
        """Экспортирует выбранные интервью в Excel файл"""
        interviews = self.get_current_filtered_interviews()

        if not interviews:
            QMessageBox.warning(self.main_window, "Ошибка", "Нет данных для экспорта")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.main_window,
            "Сохранить как",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Интервью"

            headers = [
                "ID интервью", "Дата интервью", "Анкета", "Должность",
                "ФИО кандидата", "Email", "Телефон", "Резюме", "Статус",
                "Вопрос", "Тип ответа", "Категория", "Вес вопроса",
                "Числовой ответ", "Выбранные варианты (текст)", "Текстовый ответ"
            ]

            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for interview in interviews:
                full_data = self.db.get_full_interview_data(interview['id'])
                if not full_data:
                    continue

                for row in full_data:
                    fio = f"{row['last_name']} {row['first_name']} {row['patronimic'] or ''}".strip()

                    selected_options_text = ""
                    if row['selected_options']:
                        try:
                            selected_ids = json.loads(row['selected_options'])
                            if isinstance(selected_ids, list):
                                if row['answer_type'] == "Да/Нет":
                                    selected_options_text = selected_ids[0].strip('"\'') if selected_ids else ""
                                else:
                                    ids_tuple = tuple(selected_ids)
                                    placeholders = ','.join(['%s'] * len(ids_tuple))
                                    query = f"SELECT text FROM QuestionOptions WHERE id IN ({placeholders})"
                                    options = self.db.execute_query(query, ids_tuple)
                                    selected_options_text = ", ".join([opt['text'] for opt in options])
                        except Exception as e:
                            print(f"Ошибка при обработке selected_options: {e}")
                            selected_options_text = str(row['selected_options'])

                    ws.append([
                        row['interview_id'],
                        row['interview_date'].strftime('%Y-%m-%d %H:%M'),
                        row['questionnaire_title'],
                        row['position_name'],
                        fio,
                        row['email'],
                        row['phone'],
                        row['resume_link'],
                        row['status_name'],
                        row['question_text'],
                        row['answer_type'],
                        row['category_name'],
                        row['question_weight'],
                        row['value_numeric'],
                        selected_options_text,
                        row['text_answer']
                    ])

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(self.main_window, "Успех", "Данные успешно экспортированы")
        except Exception as e:
            QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось экспортировать данные: {str(e)}")

    def get_current_filtered_interviews(self):
        """Возвращает текущий отфильтрованный список интервью"""
        query = """
        SELECT i.id, i.questionnaire_id, i.candidate_id, i.status_id, i.created_at,
               q.title as questionnaire_title,
               c.first_name, c.last_name, c.patronimic, c.resume_link,
               s.name as status_name,
               p.name as position_name
        FROM Interviews i
        JOIN Questionnaires q ON i.questionnaire_id = q.id
        JOIN Candidates c ON i.candidate_id = c.id
        JOIN Statuses s ON i.status_id = s.id
        JOIN Positions p ON q.position_id = p.id
        WHERE 1=1
        """

        params = []
        filters = self.get_current_filters()

        if self.current_user_role != 1:
            query += " AND i.interviewer_id = %s"
            params.append(self.current_user_id)

        if filters.get('status_id'):
            query += " AND i.status_id = %s"
            params.append(filters['status_id'])

        if filters.get('position_id'):
            query += " AND q.position_id = %s"
            params.append(filters['position_id'])

        if filters.get('candidate_name'):
            query += " AND (c.first_name LIKE %s OR c.last_name LIKE %s OR c.patronimic LIKE %s)"
            params.append(f'%{filters["candidate_name"]}%')
            params.append(f'%{filters["candidate_name"]}%')
            params.append(f'%{filters["candidate_name"]}%')

        if filters.get('start_date'):
            query += " AND DATE(i.created_at) >= %s"
            params.append(filters['start_date'])

        if filters.get('end_date'):
            query += " AND DATE(i.created_at) <= %s"
            params.append(filters['end_date'])

        query += " ORDER BY i.id DESC"

        return self.db.execute_query(query, params or None)

    def get_current_filters(self):
        """Возвращает текущие установленные фильтры"""
        filters = {}

        # Статус
        status_data = self.ui.statusesComboBox.currentData()
        if status_data is not None:
            filters['status_id'] = status_data

        # Должность
        position_data = self.ui.positionComboBox.currentData()
        if position_data is not None:
            filters['position_id'] = position_data

        # Имя кандидата
        candidate_name = self.ui.candidateLineEdit.text().strip()
        if candidate_name:
            filters['candidate_name'] = candidate_name

        # Даты
        start_date = self.ui.startDateEdit.date()
        if not start_date.isNull():
            filters['start_date'] = start_date.toString("yyyy-MM-dd")

        end_date = self.ui.endDateEdit.date()
        if not end_date.isNull():
            filters['end_date'] = end_date.toString("yyyy-MM-dd")

        return filters

    def import_interview_from_excel(self):
        """Импортирует интервью из Excel файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self.main_window,
            "Выберите файл для импорта",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active

            expected_headers = [
                "ID интервью", "Дата интервью", "Анкета", "Должность",
                "ФИО кандидата", "Email", "Телефон", "Резюме", "Статус",
                "Вопрос", "Тип ответа", "Категория", "Вес вопроса",
                "Числовой ответ", "Выбранные варианты (текст)", "Текстовый ответ"
            ]

            actual_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            if actual_headers != expected_headers:
                QMessageBox.warning(
                    self.main_window,
                    "Ошибка",
                    "Формат файла не соответствует ожидаемому.\n"
                    "Пожалуйста, используйте файл, экспортированный из системы."
                )
                return

            interviews_data = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                interview_id = row[0]

                if interview_id not in interviews_data:
                    status = self.db.execute_query(
                        "SELECT id FROM Statuses WHERE name = %s AND type = 'cand'",
                        (row[8],),
                        fetch_all=False
                    )

                    if not status:
                        QMessageBox.warning(
                            self.main_window,
                            "Ошибка",
                            f"Статус '{row[8]}' не найден в базе данных"
                        )
                        return

                    questionnaire = self.db.execute_query(
                        "SELECT id, position_id FROM Questionnaires WHERE title = %s",
                        (row[2],),
                        fetch_all=False
                    )

                    if not questionnaire:
                        QMessageBox.warning(
                            self.main_window,
                            "Ошибка",
                            f"Анкета '{row[2]}' не найдена в базе данных"
                        )
                        return

                    fio_parts = row[4].split()
                    last_name = fio_parts[0] if len(fio_parts) > 0 else ""
                    first_name = fio_parts[1] if len(fio_parts) > 1 else ""
                    patronimic = fio_parts[2] if len(fio_parts) > 2 else None

                    interviews_data[interview_id] = {
                        'questionnaire_id': questionnaire['id'],
                        'interviewer_id': self.current_user_id,
                        'status_id': status['id'],
                        'first_name': first_name,
                        'last_name': last_name,
                        'patronimic': patronimic,
                        'email': row[5],
                        'phone': row[6],
                        'resume_link': row[7],
                        'answers': []
                    }

                question = self.db.execute_query(
                    "SELECT id FROM Questions WHERE text = %s",
                    (row[9],),
                    fetch_all=False
                )

                if not question:
                    QMessageBox.warning(
                        self.main_window,
                        "Ошибка",
                        f"Вопрос '{row[9]}' не найден в базе данных"
                    )
                    return

                answer_type = self.db.execute_query(
                    "SELECT answer_type_id FROM Questions WHERE id = %s",
                    (question['id'],),
                    fetch_all=False
                )

                if answer_type and answer_type['answer_type_id'] == 5:
                    if row[14]:
                        answer_text = str(row[14]).lower().strip()
                        if answer_text in ['да', 'yes', 'true', '1']:
                            selected_options_ids = '["Да"]'
                            value_numeric = 1
                        else:
                            selected_options_ids = '["Нет"]'
                            value_numeric = 0
                    else:
                        selected_options_ids = '["Нет"]'
                        value_numeric = 0
                else:
                    selected_options_ids = []
                    if row[14]:
                        option_texts = [opt.strip() for opt in row[14].split(',')]
                        for text in option_texts:
                            option = self.db.execute_query(
                                "SELECT id FROM QuestionOptions WHERE text = %s AND question_id = %s",
                                (text, question['id']),
                                fetch_all=False
                            )
                            if option:
                                selected_options_ids.append(option['id'])
                    value_numeric = row[13]

                answer_data = {
                    'question_id': question['id'],
                    'value_numeric': value_numeric,
                    'selected_options': str(selected_options_ids) if selected_options_ids else None,
                    'text_answer': row[15]
                }

                interviews_data[interview_id]['answers'].append(answer_data)

            success_count = 0
            for data in interviews_data.values():
                interview_id = self.db.import_interview_data(data)
                if interview_id:
                    success_count += 1

            QMessageBox.information(
                self.main_window,
                "Импорт завершен",
                f"Успешно импортировано {success_count} из {len(interviews_data)} интервью"
            )

            self.load_candidates()

        except Exception as e:
            QMessageBox.critical(
                self.main_window,
                "Ошибка",
                f"Не удалось импортировать данные: {str(e)}"
            )
